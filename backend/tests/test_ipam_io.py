"""Tests for the IPAM import / export service layer."""

from __future__ import annotations

import csv as _csv
import io
import json

import pytest
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import hash_password
from app.models.auth import User
from app.models.ipam import IPAddress, IPBlock, IPSpace, Subnet
from app.services.ipam_io import (
    commit_import,
    export_subtree,
    parse_payload,
    preview_import,
)
from app.services.ipam_io.export import _sanitize_cell
from app.services.ipam_io.parser import ParsedPayload


async def _seed_user(db: AsyncSession) -> User:
    user = User(
        username="ioadmin",
        email="io@example.com",
        display_name="IO Admin",
        hashed_password=hash_password("x" * 10),
        is_superadmin=True,
    )
    db.add(user)
    await db.flush()
    return user


async def _seed_space(db: AsyncSession, name: str = "Test") -> IPSpace:
    space = IPSpace(name=name, description="")
    db.add(space)
    await db.flush()
    return space


# ── Parser tests ───────────────────────────────────────────────────────────────


def test_parse_csv_basic() -> None:
    data = b"network,name,vlan_id,gateway,owner\n10.0.1.0/24,Servers,10,10.0.1.1,NetOps\n"
    parsed = parse_payload(data, "import.csv", "text/csv")
    assert len(parsed.subnets) == 1
    row = parsed.subnets[0]
    assert row["network"] == "10.0.1.0/24"
    assert row["name"] == "Servers"
    assert row["vlan_id"] == 10
    assert row["gateway"] == "10.0.1.1"
    # 'owner' is unknown → custom field
    assert row["custom_fields"] == {"owner": "NetOps"}


def test_parse_json_list() -> None:
    body = json.dumps(
        [{"network": "10.0.2.0/24", "name": "App"}, {"network": "10.0.3.0/24"}]
    ).encode()
    parsed = parse_payload(body, "x.json", "application/json")
    assert [s["network"] for s in parsed.subnets] == ["10.0.2.0/24", "10.0.3.0/24"]


def test_parse_xlsx_roundtrip() -> None:
    openpyxl = pytest.importorskip("openpyxl")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "subnets"
    ws.append(["network", "name", "vlan_id"])
    ws.append(["10.5.0.0/24", "X", 42])
    buf = io.BytesIO()
    wb.save(buf)
    parsed = parse_payload(buf.getvalue(), "x.xlsx", None)
    assert len(parsed.subnets) == 1
    assert parsed.subnets[0]["network"] == "10.5.0.0/24"
    assert parsed.subnets[0]["vlan_id"] == 42


# ── Preview ────────────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_preview_creates_and_conflicts(db_session: AsyncSession) -> None:
    space = await _seed_space(db_session, "PreviewSpace")
    block = IPBlock(space_id=space.id, network="10.0.0.0/8", name="agg")
    db_session.add(block)
    await db_session.flush()
    existing = Subnet(
        space_id=space.id,
        block_id=block.id,
        network="10.0.1.0/24",
        name="already-here",
        total_ips=254,
    )
    db_session.add(existing)
    await db_session.flush()

    payload = ParsedPayload(
        subnets=[
            {"network": "10.0.1.0/24", "name": "dup"},
            {"network": "10.0.2.0/24", "name": "new"},
            {"network": "not-a-cidr"},
        ]
    )

    preview = await preview_import(db_session, payload, space_id=space.id, strategy="fail")
    assert preview.space_name == "PreviewSpace"
    created_networks = [r.network for r in preview.creates if r.kind == "subnet"]
    assert "10.0.2.0/24" in created_networks
    assert any(r.network == "10.0.1.0/24" for r in preview.conflicts)
    assert any("Invalid CIDR" in (r.reason or "") for r in preview.errors)


# ── Commit ─────────────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_commit_creates_with_auto_parent(db_session: AsyncSession) -> None:
    space = await _seed_space(db_session, "CommitSpace")
    user = await _seed_user(db_session)

    payload = ParsedPayload(
        subnets=[
            {"network": "192.168.10.0/24", "name": "lan"},
            {"network": "192.168.20.0/24", "name": "wifi", "vlan_id": 20},
        ]
    )
    result = await commit_import(
        db_session, payload, current_user=user, space_id=space.id, strategy="skip"
    )
    assert result.created_subnets == 2
    # No existing block → auto-created parents per subnet
    assert result.auto_created_blocks >= 1


@pytest.mark.asyncio
async def test_commit_overwrite(db_session: AsyncSession) -> None:
    space = await _seed_space(db_session, "OverwriteSpace")
    user = await _seed_user(db_session)
    block = IPBlock(space_id=space.id, network="10.0.0.0/8", name="agg")
    db_session.add(block)
    await db_session.flush()
    subnet = Subnet(
        space_id=space.id,
        block_id=block.id,
        network="10.0.5.0/24",
        name="old",
        description="old-desc",
        total_ips=254,
    )
    db_session.add(subnet)
    await db_session.flush()

    payload = ParsedPayload(
        subnets=[{"network": "10.0.5.0/24", "name": "new", "description": "new-desc"}]
    )
    result = await commit_import(
        db_session, payload, current_user=user, space_id=space.id, strategy="overwrite"
    )
    assert result.updated_subnets == 1
    # Flush the pending UPDATE before refresh. Session.refresh() issues a SELECT
    # without auto-flushing first, so without this it reads the unchanged row
    # and overwrites the in-memory mutation we're about to assert on.
    await db_session.flush()
    await db_session.refresh(subnet)
    assert subnet.name == "new"
    assert subnet.description == "new-desc"


# ── Export round-trip ─────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_export_csv_contains_subnets(db_session: AsyncSession) -> None:
    space = await _seed_space(db_session, "ExportSpace")
    block = IPBlock(space_id=space.id, network="172.16.0.0/12", name="b")
    db_session.add(block)
    await db_session.flush()
    db_session.add(
        Subnet(
            space_id=space.id,
            block_id=block.id,
            network="172.16.5.0/24",
            name="prod",
            total_ips=254,
        )
    )
    await db_session.flush()

    data, ctype, filename = await export_subtree(db_session, space_id=space.id, format="csv")
    assert ctype == "text/csv"
    assert filename.endswith(".csv")
    text = data.decode()
    assert "172.16.5.0/24" in text
    assert "prod" in text


@pytest.mark.asyncio
async def test_csv_roundtrip(db_session: AsyncSession) -> None:
    """Export → re-parse yields the same subnet networks."""
    space = await _seed_space(db_session, "RoundTrip")
    block = IPBlock(space_id=space.id, network="10.0.0.0/8", name="b")
    db_session.add(block)
    await db_session.flush()
    nets = ["10.0.1.0/24", "10.0.2.0/24"]
    for n in nets:
        db_session.add(
            Subnet(space_id=space.id, block_id=block.id, network=n, name=n, total_ips=254)
        )
    await db_session.flush()

    data, _, _ = await export_subtree(db_session, space_id=space.id, format="csv")
    parsed = parse_payload(data, "rt.csv", "text/csv")
    parsed_nets = sorted(str(s["network"]) for s in parsed.subnets)
    assert parsed_nets == sorted(nets)


# ── Formula-injection hardening (#523) ─────────────────────────────────────────


def test_sanitize_cell_prefixes_dangerous_values() -> None:
    # Every OWASP CSV-injection trigger gets the leading apostrophe.
    assert _sanitize_cell("=1+1") == "'=1+1"
    assert _sanitize_cell("+1") == "'+1"
    assert _sanitize_cell("-1") == "'-1"
    assert _sanitize_cell("@SUM(A1)") == "'@SUM(A1)"
    assert _sanitize_cell("\ttab") == "'\ttab"
    assert _sanitize_cell("\rcr") == "'\rcr"
    # Safe strings + non-strings pass through untouched so numeric columns
    # keep their type.
    assert _sanitize_cell("safe") == "safe"
    assert _sanitize_cell("10.0.0.0/24") == "10.0.0.0/24"
    assert _sanitize_cell(None) is None
    assert _sanitize_cell(42) == 42
    assert _sanitize_cell(True) is True


@pytest.mark.asyncio
async def test_export_csv_sanitizes_subnet_cells(db_session: AsyncSession) -> None:
    space = await _seed_space(db_session, "Inject")
    block = IPBlock(space_id=space.id, network="10.9.0.0/16", name="b")
    db_session.add(block)
    await db_session.flush()
    db_session.add(
        Subnet(
            space_id=space.id,
            block_id=block.id,
            network="10.9.1.0/24",
            name="=cmd|'/c calc'!A0",
            description="+evil",
            total_ips=254,
            custom_fields={"note": "-danger"},
        )
    )
    await db_session.flush()

    data, _, _ = await export_subtree(db_session, space_id=space.id, format="csv")
    rows = list(_csv.DictReader(io.StringIO(data.decode())))
    row = rows[0]
    assert row["name"].startswith("'=")
    assert row["description"].startswith("'+")
    assert row["note"].startswith("'-")


@pytest.mark.asyncio
async def test_export_csv_sanitizes_address_cells(db_session: AsyncSession) -> None:
    space = await _seed_space(db_session, "InjectAddr")
    block = IPBlock(space_id=space.id, network="10.10.0.0/16", name="b")
    db_session.add(block)
    await db_session.flush()
    subnet = Subnet(
        space_id=space.id, block_id=block.id, network="10.10.1.0/24", name="s", total_ips=254
    )
    db_session.add(subnet)
    await db_session.flush()
    db_session.add(
        IPAddress(
            subnet_id=subnet.id,
            address="10.10.1.5",
            status="allocated",
            hostname="@SUM(1)",
            description="=danger",
            custom_fields={"owner": "-x"},
        )
    )
    await db_session.flush()

    # include_addresses on a subnet scope emits the addresses-only CSV.
    data, _, _ = await export_subtree(
        db_session, subnet_id=subnet.id, format="csv", include_addresses=True
    )
    rows = list(_csv.DictReader(io.StringIO(data.decode())))
    row = rows[0]
    assert row["hostname"].startswith("'@")
    assert row["description"].startswith("'=")
    assert row["owner"].startswith("'-")


@pytest.mark.asyncio
async def test_export_xlsx_sanitizes_cells(db_session: AsyncSession) -> None:
    from openpyxl import load_workbook

    space = await _seed_space(db_session, "InjectXlsx")
    block = IPBlock(space_id=space.id, network="10.11.0.0/16", name="=blkname")
    db_session.add(block)
    await db_session.flush()
    subnet = Subnet(
        space_id=space.id,
        block_id=block.id,
        network="10.11.1.0/24",
        name="=cmd()",
        total_ips=254,
    )
    db_session.add(subnet)
    await db_session.flush()
    db_session.add(
        IPAddress(
            subnet_id=subnet.id,
            address="10.11.1.5",
            status="allocated",
            hostname="+evil",
        )
    )
    await db_session.flush()

    data, _, _ = await export_subtree(
        db_session, space_id=space.id, format="xlsx", include_addresses=True
    )
    wb = load_workbook(io.BytesIO(data))

    subnets_ws = wb["subnets"]
    name_col = [c.value for c in subnets_ws[1]].index("name")
    assert str(subnets_ws[2][name_col].value).startswith("'=")

    blocks_ws = wb["blocks"]
    b_name_col = [c.value for c in blocks_ws[1]].index("name")
    assert str(blocks_ws[2][b_name_col].value).startswith("'=")

    addr_ws = wb["addresses"]
    host_col = [c.value for c in addr_ws[1]].index("hostname")
    assert str(addr_ws[2][host_col].value).startswith("'+")
