"""IPAM export — CSV, JSON, XLSX serialisation of a subtree.

``export_subtree`` returns a ``(bytes, content_type, filename)`` tuple —
the router streams this back to the client.  The caller selects the
scope (an IP space, a block, or a single subnet); the exporter walks
down the hierarchy and collects everything below the root.
"""

from __future__ import annotations

import csv
import io
import ipaddress
import json
import uuid
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Literal

from fastapi import HTTPException, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.ipam import IPAddress, IPBlock, IPSpace, Subnet

Format = Literal["csv", "json", "xlsx"]

SUBNET_COLUMNS = [
    "space",
    "block",
    "network",
    "name",
    "description",
    "gateway",
    "vlan_id",
    "vxlan_id",
    "status",
    "total_ips",
    "allocated_ips",
    "utilization_percent",
]


@dataclass
class ExportBundle:
    space: dict[str, Any] | None
    blocks: list[dict[str, Any]]
    subnets: list[dict[str, Any]]
    addresses: list[dict[str, Any]]


def _dt(value: datetime | None) -> str | None:
    return value.isoformat() if value else None


async def _collect(
    db: AsyncSession,
    *,
    space_id: uuid.UUID | None,
    block_id: uuid.UUID | None,
    subnet_id: uuid.UUID | None,
    include_addresses: bool,
) -> ExportBundle:
    if not any([space_id, block_id, subnet_id]):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Export requires one of space_id, block_id, or subnet_id",
        )

    # Resolve scope
    space: IPSpace | None
    blocks_raw: list[IPBlock]
    subnets: list[Subnet]

    if subnet_id:
        subnet = await db.get(Subnet, subnet_id)
        if subnet is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Subnet not found")
        space = await db.get(IPSpace, subnet.space_id)
        subnets = [subnet]
        blocks_raw = []
        block = await db.get(IPBlock, subnet.block_id) if subnet.block_id else None
        if block:
            blocks_raw.append(block)
    elif block_id:
        block = await db.get(IPBlock, block_id)
        if block is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Block not found")
        space = await db.get(IPSpace, block.space_id)
        all_blocks_res = await db.execute(select(IPBlock).where(IPBlock.space_id == block.space_id))
        all_blocks = list(all_blocks_res.scalars().all())
        block_net = ipaddress.ip_network(str(block.network), strict=False)
        blocks_raw = [
            b for b in all_blocks if _is_subnet_of(str(b.network), block_net) or b.id == block.id
        ]
        block_ids = {b.id for b in blocks_raw}
        subnets_res = await db.execute(select(Subnet).where(Subnet.block_id.in_(block_ids)))
        subnets = list(subnets_res.scalars().all())
    else:
        space = await db.get(IPSpace, space_id)
        if space is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="IP space not found")
        blocks_res = await db.execute(select(IPBlock).where(IPBlock.space_id == space.id))
        blocks_raw = list(blocks_res.scalars().all())
        subnets_res = await db.execute(select(Subnet).where(Subnet.space_id == space.id))
        subnets = list(subnets_res.scalars().all())

    if space is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Parent IP space not found"
        )

    block_by_id = {b.id: b for b in blocks_raw}
    space_name = space.name

    blocks_payload = [
        {
            "id": str(b.id),
            "parent_block_id": str(b.parent_block_id) if b.parent_block_id else None,
            "network": str(b.network),
            "name": b.name,
            "description": b.description,
            "utilization_percent": b.utilization_percent,
            "custom_fields": b.custom_fields or {},
        }
        for b in blocks_raw
    ]
    subnets_payload = [
        {
            "space": space_name,
            "block": str(block_by_id[s.block_id].network) if s.block_id in block_by_id else None,
            "network": str(s.network),
            "name": s.name,
            "description": s.description,
            "gateway": str(s.gateway) if s.gateway else None,
            "vlan_id": s.vlan_id,
            "vxlan_id": s.vxlan_id,
            "status": s.status,
            "total_ips": s.total_ips,
            "allocated_ips": s.allocated_ips,
            "utilization_percent": s.utilization_percent,
            "custom_fields": s.custom_fields or {},
        }
        for s in subnets
    ]

    addresses_payload: list[dict[str, Any]] = []
    if include_addresses and subnets:
        subnet_ids = [s.id for s in subnets]
        addrs_res = await db.execute(select(IPAddress).where(IPAddress.subnet_id.in_(subnet_ids)))
        subnet_nets = {s.id: str(s.network) for s in subnets}
        for a in addrs_res.scalars().all():
            addresses_payload.append(
                {
                    "subnet": subnet_nets.get(a.subnet_id),
                    "address": str(a.address),
                    "status": a.status,
                    "hostname": a.hostname,
                    "fqdn": a.fqdn,
                    "mac_address": str(a.mac_address) if a.mac_address else None,
                    "description": a.description,
                    "last_seen_at": _dt(a.last_seen_at),
                    "custom_fields": a.custom_fields or {},
                }
            )

    return ExportBundle(
        space={"id": str(space.id), "name": space.name, "description": space.description},
        blocks=blocks_payload,
        subnets=subnets_payload,
        addresses=addresses_payload,
    )


def _is_subnet_of(
    child: str,
    parent_net: ipaddress.IPv4Network | ipaddress.IPv6Network,
) -> bool:
    try:
        child_net = ipaddress.ip_network(child, strict=False)
    except ValueError:
        return False
    if child_net.version != parent_net.version:
        return False
    return child_net.subnet_of(parent_net)  # type: ignore[arg-type]


# ── Serialisers ────────────────────────────────────────────────────────────────


def _to_csv(bundle: ExportBundle) -> bytes:
    buf = io.StringIO()
    # Collect any custom field keys so each gets its own column.
    cf_keys: list[str] = []
    seen: set[str] = set()
    for row in bundle.subnets:
        for k in (row.get("custom_fields") or {}).keys():
            if k not in seen:
                seen.add(k)
                cf_keys.append(k)
    headers = SUBNET_COLUMNS + cf_keys
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in bundle.subnets:
        out = {k: row.get(k) for k in SUBNET_COLUMNS}
        for k in cf_keys:
            out[k] = (row.get("custom_fields") or {}).get(k)
        writer.writerow(out)
    return buf.getvalue().encode("utf-8")


def _to_json(bundle: ExportBundle) -> bytes:
    payload = {
        "space": bundle.space,
        "blocks": bundle.blocks,
        "subnets": bundle.subnets,
        "addresses": bundle.addresses,
    }
    return json.dumps(payload, indent=2, default=str).encode("utf-8")


def _to_xlsx(bundle: ExportBundle) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl is required for XLSX export but is not installed",
        ) from exc

    wb = Workbook()
    # Subnets sheet
    ws = wb.active
    ws.title = "subnets"
    cf_keys: list[str] = []
    seen: set[str] = set()
    for row in bundle.subnets:
        for k in (row.get("custom_fields") or {}).keys():
            if k not in seen:
                seen.add(k)
                cf_keys.append(k)
    headers = SUBNET_COLUMNS + cf_keys
    ws.append(headers)
    for row in bundle.subnets:
        base = [row.get(k) for k in SUBNET_COLUMNS]
        cf = row.get("custom_fields") or {}
        ws.append(base + [cf.get(k) for k in cf_keys])

    # Blocks sheet
    blocks_ws = wb.create_sheet("blocks")
    blocks_ws.append(
        ["id", "parent_block_id", "network", "name", "description", "utilization_percent"]
    )
    for b in bundle.blocks:
        blocks_ws.append(
            [
                b.get("id"),
                b.get("parent_block_id"),
                b.get("network"),
                b.get("name"),
                b.get("description"),
                b.get("utilization_percent"),
            ]
        )

    # Addresses sheet
    if bundle.addresses:
        addr_ws = wb.create_sheet("addresses")
        addr_ws.append(
            [
                "subnet",
                "address",
                "status",
                "hostname",
                "fqdn",
                "mac_address",
                "description",
                "last_seen_at",
            ]
        )
        for a in bundle.addresses:
            addr_ws.append(
                [
                    a.get("subnet"),
                    a.get("address"),
                    a.get("status"),
                    a.get("hostname"),
                    a.get("fqdn"),
                    a.get("mac_address"),
                    a.get("description"),
                    a.get("last_seen_at"),
                ]
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Public API ─────────────────────────────────────────────────────────────────


async def export_subtree(
    db: AsyncSession,
    *,
    space_id: uuid.UUID | None = None,
    block_id: uuid.UUID | None = None,
    subnet_id: uuid.UUID | None = None,
    format: Format = "csv",
    include_addresses: bool = False,
) -> tuple[bytes, str, str]:
    bundle = await _collect(
        db,
        space_id=space_id,
        block_id=block_id,
        subnet_id=subnet_id,
        include_addresses=include_addresses,
    )
    scope_name = bundle.space["name"] if bundle.space else "export"
    safe_name = str(scope_name).replace("/", "_").replace(" ", "_")
    if format == "csv":
        return _to_csv(bundle), "text/csv", f"ipam-{safe_name}.csv"
    if format == "json":
        return _to_json(bundle), "application/json", f"ipam-{safe_name}.json"
    if format == "xlsx":
        return (
            _to_xlsx(bundle),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"ipam-{safe_name}.xlsx",
        )
    raise HTTPException(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        detail=f"Unsupported export format: {format}",
    )
