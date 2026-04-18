"""Format parsers for IPAM import payloads.

Input formats supported:
- CSV (first row is header; columns: network, name, gateway, vlan_id,
  description, space, block, plus any custom field columns).
- JSON (either a plain list of subnet objects, or an object with
  ``spaces``, ``blocks``, ``subnets``, ``addresses`` keys for a full
  hierarchical dump as produced by the exporter).
- XLSX (workbook with a ``subnets`` sheet at minimum; optional
  ``spaces``, ``blocks``, ``addresses`` sheets).

All parsers return a :class:`ParsedPayload` — a format-neutral intermediate
that the importer consumes.
"""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass, field
from typing import Any

from fastapi import HTTPException, status


@dataclass
class ParsedPayload:
    """Normalised import payload — lists of dict rows, one per resource type."""

    spaces: list[dict[str, Any]] = field(default_factory=list)
    blocks: list[dict[str, Any]] = field(default_factory=list)
    subnets: list[dict[str, Any]] = field(default_factory=list)
    addresses: list[dict[str, Any]] = field(default_factory=list)


_KNOWN_SUBNET_COLUMNS = {
    "network",
    "name",
    "gateway",
    "vlan_id",
    "vxlan_id",
    "description",
    "status",
    "domain_name",
    "space",
    "space_name",
    "block",
    "block_network",
}


_KNOWN_ADDRESS_COLUMNS = {
    "address",
    "ip",
    "ip_address",
    "hostname",
    "fqdn",
    "mac",
    "mac_address",
    "description",
    "status",
    "tags",
    "subnet",
}


def _coerce_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _row_to_subnet(row: dict[str, Any]) -> dict[str, Any]:
    """Coerce a generic row dict into a subnet dict with custom_fields."""
    out: dict[str, Any] = {}
    custom: dict[str, Any] = {}
    for raw_key, value in row.items():
        if raw_key is None:
            continue
        key = raw_key.strip()
        if not key:
            continue
        norm = key.lower()
        if value == "":
            value = None
        if norm in _KNOWN_SUBNET_COLUMNS:
            out[norm] = value
        elif norm in {"vlan", "vlan id"}:
            out["vlan_id"] = value
        elif norm in {"vxlan", "vxlan id"}:
            out["vxlan_id"] = value
        else:
            # Everything unrecognised becomes a custom field.
            custom[key] = value
    if custom:
        out["custom_fields"] = custom
    # Normalise vlan_id
    if "vlan_id" in out:
        out["vlan_id"] = _coerce_int(out["vlan_id"])
    if "vxlan_id" in out:
        out["vxlan_id"] = _coerce_int(out["vxlan_id"])
    return out


def _row_to_address(row: dict[str, Any]) -> dict[str, Any]:
    """Coerce a generic row dict into an IPAddress dict with custom_fields.

    Recognises ``address`` / ``ip`` / ``ip_address`` as the canonical
    address column so exports from other DDI tools map cleanly. Tags
    and custom_fields are passed through unchanged when already dicts
    and preserved verbatim on pass-through exports from our own
    exporter. Unrecognised columns become ``custom_fields`` entries —
    same pattern as the subnet coercer — so a CSV like
    ``address,hostname,department,cost_center`` Just Works.
    """
    out: dict[str, Any] = {}
    custom: dict[str, Any] = {}
    for raw_key, value in row.items():
        if raw_key is None:
            continue
        key = raw_key.strip()
        if not key:
            continue
        norm = key.lower().replace(" ", "_")
        if value == "":
            value = None
        if norm in ("ip", "ip_address"):
            out["address"] = value
        elif norm == "mac":
            out["mac_address"] = value
        elif norm in _KNOWN_ADDRESS_COLUMNS:
            out[norm] = value
        elif norm in ("custom_fields", "customfields"):
            # Pass-through from our exporter — value is already a dict.
            if isinstance(value, dict):
                custom.update(value)
        else:
            custom[key] = value
    if custom:
        out["custom_fields"] = custom
    return out


def _csv_is_address_payload(headers: list[str]) -> bool:
    """Detect an address-upload CSV by header row.

    Subnet imports must have ``network``; address imports must have
    ``address`` / ``ip`` / ``ip_address``. When both are present (the
    ``network`` column wins and the row is treated as a subnet) — the
    two formats shouldn't mix in a single file.
    """
    norm = {h.strip().lower() for h in headers if h}
    if "network" in norm:
        return False
    return bool({"address", "ip", "ip_address"} & norm)


def parse_csv(data: bytes) -> ParsedPayload:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    payload = ParsedPayload()
    is_addresses = _csv_is_address_payload(list(reader.fieldnames or []))
    for row in reader:
        if not any((v or "").strip() for v in row.values() if isinstance(v, str)):
            continue
        if is_addresses:
            if not any(row.get(k) for k in ("address", "ip", "ip_address")):
                continue
            payload.addresses.append(_row_to_address(row))
        else:
            if not row.get("network") and not row.get("Network"):
                continue
            payload.subnets.append(_row_to_subnet(row))
    return payload


def parse_json(data: bytes) -> ParsedPayload:
    try:
        decoded = json.loads(data.decode("utf-8-sig"))
    except (json.JSONDecodeError, UnicodeDecodeError) as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Invalid JSON payload: {exc}",
        ) from exc

    payload = ParsedPayload()
    if isinstance(decoded, list):
        # List form is ambiguous — sniff the first row for an address key
        # and route accordingly. Mixed lists aren't supported (don't do that).
        first = next((x for x in decoded if isinstance(x, dict)), None)
        is_addresses = first is not None and any(
            k.lower() in ("address", "ip", "ip_address") for k in first.keys()
        )
        for item in decoded:
            if not isinstance(item, dict):
                continue
            if is_addresses:
                payload.addresses.append(_row_to_address(item))
            else:
                payload.subnets.append(_row_to_subnet(item))
        return payload

    if isinstance(decoded, dict):
        for key in ("spaces", "blocks", "subnets", "addresses"):
            items = decoded.get(key) or []
            if not isinstance(items, list):
                continue
            for item in items:
                if not isinstance(item, dict):
                    continue
                if key == "subnets":
                    payload.subnets.append(_row_to_subnet(item))
                elif key == "addresses":
                    payload.addresses.append(_row_to_address(item))
                else:
                    getattr(payload, key).append(dict(item))
        return payload

    raise HTTPException(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        detail="JSON payload must be a list of subnets or an object with "
        "spaces/blocks/subnets/addresses keys",
    )


def parse_xlsx(data: bytes) -> ParsedPayload:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - runtime dep guard
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl is required for XLSX import but is not installed",
        ) from exc

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Invalid XLSX file: {exc}",
        ) from exc

    payload = ParsedPayload()

    def _sheet_rows(sheet_name: str) -> list[dict[str, Any]]:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return []
        headers = [str(h).strip() if h is not None else "" for h in header]
        out: list[dict[str, Any]] = []
        for raw in rows:
            if raw is None or all(c is None or c == "" for c in raw):
                continue
            row = {headers[i]: raw[i] for i in range(min(len(headers), len(raw))) if headers[i]}
            out.append(row)
        return out

    # Primary sheet routing — a workbook with only an ``addresses`` sheet
    # (or ``subnets``) is the common shape for a per-resource upload; a
    # workbook with both sheets is our own exporter's output and all
    # sheets are loaded.
    if "subnets" in wb.sheetnames:
        for row in _sheet_rows("subnets"):
            if not row.get("network") and not row.get("Network"):
                continue
            payload.subnets.append(_row_to_subnet(row))
    elif "addresses" not in wb.sheetnames and wb.sheetnames:
        # Single-sheet workbook with neither well-known name — sniff the
        # first row's headers to decide whether it's subnets or addresses.
        first_sheet = wb.sheetnames[0]
        ws = wb[first_sheet]
        headers_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        first_row = next(iter(headers_iter), ())
        headers_norm = {(str(h).strip().lower() if h is not None else "") for h in first_row}
        is_addresses = bool({"address", "ip", "ip_address"} & headers_norm) and (
            "network" not in headers_norm
        )
        for row in _sheet_rows(first_sheet):
            if is_addresses:
                if not any(row.get(k) for k in ("address", "ip", "ip_address")):
                    continue
                payload.addresses.append(_row_to_address(row))
            else:
                if not row.get("network") and not row.get("Network"):
                    continue
                payload.subnets.append(_row_to_subnet(row))

    if "addresses" in wb.sheetnames:
        for row in _sheet_rows("addresses"):
            if not any(row.get(k) for k in ("address", "ip", "ip_address")):
                continue
            payload.addresses.append(_row_to_address(row))

    for key in ("spaces", "blocks"):
        for row in _sheet_rows(key):
            getattr(payload, key).append(dict(row))

    wb.close()
    return payload


def parse_payload(data: bytes, filename: str, content_type: str | None) -> ParsedPayload:
    """Dispatch to the format parser based on the file extension / MIME type."""
    name = (filename or "").lower()
    ctype = (content_type or "").lower()
    if name.endswith(".csv") or "csv" in ctype:
        return parse_csv(data)
    if name.endswith(".json") or "json" in ctype:
        return parse_json(data)
    if name.endswith(".xlsx") or "spreadsheetml" in ctype or "officedocument" in ctype:
        return parse_xlsx(data)
    raise HTTPException(
        status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
        detail=f"Unsupported import format: {filename or content_type or 'unknown'}. "
        "Use CSV, JSON, or XLSX.",
    )
